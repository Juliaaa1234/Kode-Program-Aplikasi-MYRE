from datetime import datetime, date
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook, load_workbook
import os

class DatabaseManager:
    def __init__(self, filename='database.xlsx'):
        self.filename = filename
        self.init_database()
    
    def init_database(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            
            ws_users = wb.active
            ws_users.title = "users"
            ws_users.append(['username', 'password', 'nama', 'nim', 'ipk'])
            
            ws_jadwal = wb.create_sheet("jadwal")
            ws_jadwal.append(['username', 'hari', 'mata_kuliah', 'waktu', 'ruangan', 'dosen'])
            
            ws_todo = wb.create_sheet("todo")
            ws_todo.append(['username', 'aktivitas', 'status', 'created_date'])
            
            ws_tugas = wb.create_sheet("tugas")
            ws_tugas.append(['username', 'nama_tugas', 'mata_kuliah', 'deadline', 'status'])
            
            ws_kalender = wb.create_sheet("kalender")
            ws_kalender.append(['username', 'tanggal', 'aktivitas', 'waktu'])
            
            ws_profile = wb.create_sheet("profile")
            ws_profile.append(['username', 'nama', 'nim', 'ipk'])
            
            wb.save(self.filename)
    
    def register_user(self, username: str, password: str) -> bool:
        wb = load_workbook(self.filename)
        ws = wb["users"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username:
                wb.close()
                return False

        ws.append([username, password, '', '', ''])

        ws_profile = wb["profile"]
        ws_profile.append([username, '', '', ''])
        
        wb.save(self.filename)
        wb.close()
        return True
    
    def login_user(self, username: str, password: str) -> bool:
        wb = load_workbook(self.filename)
        ws = wb["users"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username and row[1] == password:
                wb.close()
                return True
        wb.close()
        return False
    
    def update_profile(self, username: str, nama: str, nim: str, ipk: str):
        wb = load_workbook(self.filename)
        ws = wb["profile"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == username:
                row[1].value = nama
                row[2].value = nim
                row[3].value = ipk
                break

        ws_users = wb["users"]
        for row in ws_users.iter_rows(min_row=2):
            if row[0].value == username:
                row[2].value = nama
                row[3].value = nim
                row[4].value = ipk
                break
        
        wb.save(self.filename)
        wb.close()
    
    def get_profile(self, username: str) -> Dict[str, str]:
        wb = load_workbook(self.filename)
        ws = wb["profile"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username:
                wb.close()
                return {
                    'username': row[0],
                    'nama': row[1] or '',
                    'nim': row[2] or '',
                    'ipk': row[3] or ''
                }
        wb.close()
        return {}

    def add_todo(self, username: str, aktivitas: str):
        wb = load_workbook(self.filename)
        ws = wb["todo"]
        
        ws.append([username, aktivitas, 'pending', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        wb.save(self.filename)
        wb.close()
    
    def get_todos(self, username: str) -> List[Dict]:
        wb = load_workbook(self.filename)
        ws = wb["todo"]
        
        todos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username:
                todos.append({
                    'aktivitas': row[1],
                    'status': row[2],
                    'created_date': row[3]
                })
        
        wb.close()
        return todos
    
    def update_todo_status(self, username: str, aktivitas: str, status: str):
        wb = load_workbook(self.filename)
        ws = wb["todo"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == username and row[1].value == aktivitas:
                row[2].value = status
                break

        wb.save(self.filename)
        wb.close()

    def update_todo_aktivitas(self, username: str, old_aktivitas: str, new_aktivitas: str):
        wb = load_workbook(self.filename)
        ws = wb["todo"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == username and row[1].value == old_aktivitas:
                row[1].value = new_aktivitas
                break

        wb.save(self.filename)
        wb.close()

    def delete_todo(self, username: str, aktivitas: str):
        wb = load_workbook(self.filename)
        ws = wb["todo"]

        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == username and row[1].value == aktivitas:
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        wb.save(self.filename)
        wb.close()

    def get_jadwal_hari_ini(self, username: str, hari: str) -> List[Dict]:
        wb = load_workbook(self.filename)
        ws = wb["jadwal"]

        jadwal_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username and row[1].lower() == hari.lower():
                jadwal_list.append({
                    'hari': row[1],
                    'mata_kuliah': row[2],
                    'waktu': row[3],
                    'ruangan': row[4],
                    'dosen': row[5]
                })

        wb.close()
        return jadwal_list
    
    def add_tugas(self, username: str, nama_tugas: str, mata_kuliah: str, deadline: str):
        wb = load_workbook(self.filename)
        ws = wb["tugas"]
        
        ws.append([username, nama_tugas, mata_kuliah, deadline, 'pending'])
        wb.save(self.filename)
        wb.close()
    
    def get_tugas_by_user(self, username: str) -> List[Dict]:
        wb = load_workbook(self.filename)
        ws = wb["tugas"]

        tugas_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username:
                tugas_list.append({
                    'nama_tugas': row[1],
                    'mata_kuliah': row[2],
                    'deadline': row[3],
                    'status': row[4]
                })

        wb.close()
        return tugas_list

    def update_tugas(self, username: str, old_nama_tugas: str, new_nama_tugas: str, new_mata_kuliah: str, new_deadline: str):
        wb = load_workbook(self.filename)
        ws = wb["tugas"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == username and row[1].value == old_nama_tugas:
                row[1].value = new_nama_tugas
                row[2].value = new_mata_kuliah
                row[3].value = new_deadline
                break

        wb.save(self.filename)
        wb.close()

    def delete_tugas(self, username: str, nama_tugas: str):
        wb = load_workbook(self.filename)
        ws = wb["tugas"]

        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == username and row[1].value == nama_tugas:
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        wb.save(self.filename)
        wb.close()

    def update_tugas_status(self, username: str, nama_tugas: str, status: str):
        wb = load_workbook(self.filename)
        ws = wb["tugas"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == username and row[1].value == nama_tugas:
                row[4].value = status
                break

        wb.save(self.filename)
        wb.close()

    def add_jadwal(self, username: str, hari: str, mata_kuliah: str, 
                   waktu: str, ruangan: str, dosen: str):
        wb = load_workbook(self.filename)
        ws = wb["jadwal"]
        
        ws.append([username, hari, mata_kuliah, waktu, ruangan, dosen])
        wb.save(self.filename)
        wb.close()
    
    def get_jadwal_by_user(self, username: str) -> List[Dict]:
        wb = load_workbook(self.filename)
        ws = wb["jadwal"]
        
        jadwal_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username:
                jadwal_list.append({
                    'hari': row[1],
                    'mata_kuliah': row[2],
                    'waktu': row[3],
                    'ruangan': row[4],
                    'dosen': row[5]
                })
        
        wb.close()
        return jadwal_list

    def update_jadwal(self, username: str, old_mata_kuliah: str, 
                      new_hari: str, new_mata_kuliah: str, 
                      new_waktu: str, new_ruangan: str, new_dosen: str):
        
        wb = load_workbook(self.filename)
        ws = wb["jadwal"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == username and row[2].value == old_mata_kuliah:
                row[1].value = new_hari
                row[2].value = new_mata_kuliah
                row[3].value = new_waktu
                row[4].value = new_ruangan
                row[5].value = new_dosen
                break

        wb.save(self.filename)
        wb.close()

    def delete_jadwal(self, username: str, mata_kuliah: str):
        wb = load_workbook(self.filename)
        ws = wb["jadwal"]

        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == username and row[2].value == mata_kuliah:
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        wb.save(self.filename)
        wb.close()
    
    def add_kalender_aktivitas(self, username: str, tanggal: str, aktivitas: str, waktu: str = ''):
        wb = load_workbook(self.filename)
        ws = wb["kalender"]
        
        ws.append([username, tanggal, aktivitas, waktu])
        wb.save(self.filename)
        wb.close()
    
    def get_kalender_by_month(self, username: str, bulan: int, tahun: int = None) -> Dict[str, List]:
        if tahun is None:
            tahun = datetime.now().year

        wb = load_workbook(self.filename)
        ws = wb["kalender"]

        aktivitas_per_tanggal = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == username:
                try:
                    tanggal_obj = datetime.strptime(str(row[1]), '%Y-%m-%d')
                    if tanggal_obj.month == bulan and tanggal_obj.year == tahun:
                        tanggal_str = tanggal_obj.strftime('%Y-%m-%d')
                        if tanggal_str not in aktivitas_per_tanggal:
                            aktivitas_per_tanggal[tanggal_str] = []

                        aktivitas_per_tanggal[tanggal_str].append({
                            'aktivitas': row[2],
                            'waktu': row[3] or ''
                        })
                except:
                    continue

        wb.close()
        return aktivitas_per_tanggal

    def update_kalender_aktivitas(self, username: str, tanggal: str, old_aktivitas: str, new_aktivitas: str, waktu: str = ''):
        wb = load_workbook(self.filename)
        ws = wb["kalender"]

        for row in ws.iter_rows(min_row=2):
            if (row[0].value == username and row[1].value == tanggal and row[2].value == old_aktivitas):
                row[2].value = new_aktivitas
                row[3].value = waktu
                break

        wb.save(self.filename)
        wb.close()

    def delete_kalender_aktivitas(self, username: str, tanggal: str, aktivitas: str):
        wb = load_workbook(self.filename)
        ws = wb["kalender"]

        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if (row[0].value == username and row[1].value == tanggal and row[2].value == aktivitas):
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        wb.save(self.filename)
        wb.close()
